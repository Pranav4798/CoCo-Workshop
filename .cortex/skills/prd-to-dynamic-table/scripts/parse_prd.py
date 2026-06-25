# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Parse a PRD-style XLSX workbook into structured JSON for Dynamic Table planning.

Heuristically identifies sheets as: source_onboarding, column_mappings, or business_rules
based on header patterns. Outputs JSON to stdout.
"""
import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

# Header keywords used to classify sheets
SOURCE_ONBOARDING_SIGNALS = {"source system", "erp platform", "region", "go-live", "priority", "requested by"}
COLUMN_MAPPING_SIGNALS = {"source column", "target column", "silver target", "source type", "target type", "transform"}
BUSINESS_RULES_SIGNALS = {"rule id", "rule description", "category", "decision"}


def normalize_header(val: str) -> str:
    """Lowercase, strip whitespace for matching."""
    return str(val).strip().lower() if val else ""


def classify_sheet(headers: list[str]) -> str | None:
    """Return sheet classification based on header overlap with known patterns."""
    normed = {normalize_header(h) for h in headers if h}

    scores = {
        "source_onboarding": len(normed & SOURCE_ONBOARDING_SIGNALS),
        "column_mappings": len(normed & COLUMN_MAPPING_SIGNALS),
        "business_rules": len(normed & BUSINESS_RULES_SIGNALS),
    }

    best = max(scores, key=scores.get)
    if scores[best] >= 2:
        return best
    return None


def extract_rows(ws) -> tuple[list[str], list[dict]]:
    """Extract header row and data rows from a worksheet, skipping empty rows."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    # Find header row (first row with at least 3 non-empty cells)
    header_idx = None
    for i, row in enumerate(rows):
        non_empty = sum(1 for c in row if c is not None and str(c).strip())
        if non_empty >= 3:
            header_idx = i
            break

    if header_idx is None:
        return [], []

    headers = [str(c).strip() if c else f"col_{j}" for j, c in enumerate(rows[header_idx])]
    data = []

    for row in rows[header_idx + 1:]:
        # Skip fully empty rows
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        record = {}
        for j, val in enumerate(row):
            if j < len(headers):
                # Convert non-string types to string for JSON serialization
                if val is None:
                    record[headers[j]] = None
                elif hasattr(val, "isoformat"):
                    record[headers[j]] = val.isoformat()
                else:
                    record[headers[j]] = val
        data.append(record)

    return headers, data


def parse_workbook(path: Path) -> dict:
    """Parse all sheets in the workbook and return classified data."""
    wb = load_workbook(path, read_only=True, data_only=True)
    result = {
        "source_onboarding": [],
        "column_mappings": [],
        "business_rules": [],
        "unclassified_sheets": [],
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers, data = extract_rows(ws)

        if not headers or not data:
            continue

        classification = classify_sheet(headers)
        if classification:
            result[classification].extend(data)
        else:
            result["unclassified_sheets"].append({
                "sheet_name": sheet_name,
                "headers": headers,
                "row_count": len(data),
            })

    wb.close()

    # Summary for quick validation
    result["_summary"] = {
        "source_onboarding_rows": len(result["source_onboarding"]),
        "column_mapping_rows": len(result["column_mappings"]),
        "business_rules_rows": len(result["business_rules"]),
        "unclassified_sheet_count": len(result["unclassified_sheets"]),
    }

    return result


def main():
    parser = argparse.ArgumentParser(
        description="Parse a PRD XLSX workbook into structured JSON for DT planning."
    )
    parser.add_argument(
        "--input", "-i",
        required=True,
        help="Path to the XLSX requirements file",
    )
    args = parser.parse_args()

    path = Path(args.input)
    if not path.exists():
        print(f"Error: File not found: {path}", file=sys.stderr)
        sys.exit(1)
    if path.suffix.lower() not in (".xlsx", ".xls"):
        print(f"Error: Expected .xlsx file, got: {path.suffix}", file=sys.stderr)
        sys.exit(1)

    try:
        result = parse_workbook(path)
    except Exception as e:
        print(f"Error parsing workbook: {e}", file=sys.stderr)
        sys.exit(1)

    json.dump(result, sys.stdout, indent=2, default=str)
    print()  # trailing newline


if __name__ == "__main__":
    main()
